"""
PROPOSTA = Criar um codigo capaz de transformar varios arquivos csv de uma pasta em xlsx e uni-los em um único arquivo xlsx com várias abas, uma para cada arquivo.
ETAPAS =
1 = Transformar os arquivos de csv para xlsx automaticamente
2 = Filtrar cada xlsx individualmente, selecionando na coluna clinica, somente as 47, 73, 87, 104
3 = Remover as colunas com nomes e informaçoes pessoais do paciente
4 = Unir todas planilhas em uma só pasta de trabalho excel, cada um com seu nome respectivo
5 = Salvar o arquivo
"""
from openpyxl import Workbook  #IMPORTAÇÃO DA CLASSE WORKBOOK
import os
import pandas as pd

pasta = r"C:\Users\allisson.avila\Documents\GitHub\Python-Study\Estagio\VariasPlanilhasEmAbas\arqUnir"

# Cria o arquivo de saída
pastatrabalho = Workbook()

# Remove a aba padrão criada automaticamente
aba_padrao = pastatrabalho.active
pastatrabalho.remove(aba_padrao)

# Contador para nomear abas
contador = 1

# Loop pelos arquivos da pasta
for arq in os.listdir(pasta):
    if arq.endswith(".xlsx"):
        caminho = os.path.join(pasta, arq)  # Monta o caminho completo
        df = pd.read_excel(caminho)         # Lê o Excel

        # Cria uma nova aba com nome único
        aba = pastatrabalho.create_sheet(title=f"mes_{contador}")

        # Escreve o cabeçalho
        for col_num, coluna in enumerate(df.columns, 1):
            aba.cell(row=1, column=col_num, value=coluna)

        # Escreve os dados
        for row_num, row in enumerate(df.values, 2):
            for col_num, valor in enumerate(row, 1):
                aba.cell(row=row_num, column=col_num, value=valor)

        contador += 1

# Salva o arquivo final
pastatrabalho.save("todas_planilhas.xlsx")



