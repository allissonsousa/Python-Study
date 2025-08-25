"""
PROPOSTA = Criar um codigo capaz de transformar varios arquivos csv de uma pasta em xlsx e uni-los em um único arquivo xlsx com várias abas, uma para cada arquivo.
ETAPAS =
1 = Transformar os arquivos de csv para xlsx automaticamente
2 = Filtrar cada xlsx individualmente, selecionando na coluna clinica, somente as 47, 73, 87, 104
3 = Remover as colunas com nomes e informaçoes pessoais do paciente
4 = Unir todas planilhas em uma só pasta de trabalho excel, cada um com seu nome respectivo
5 = Salvar o arquivo
"""
from openpyxl import Workbook  #IMPORTAÇÃO DA CLASSE WORKBOOK
import os
import pandas as pd

pasta = r"C:\Users\allisson.avila\Documents\GitHub\Python-Study\Testes\UniaoDeAbasXLSX\arqUnir"

pastatrabalho = Workbook()
# Lista para armazenar os DataFrames

# Loop pelos arquivos na pasta
contador = 1
for arq in os.listdir(pasta):
    if arq.endswith(".csv"):
        caminho = os.path.join(pasta, arq)     #monta o caminho completo
        df = pd.read_csv(caminho, encoding="ISO-8859-1")  #tenta abrir o csv
       # Adiciona a primeira aba (geralmente 'Sheet') e a seleciona

        aba = pastatrabalho.create_sheet(title = f"mes {contador}")

        for col_num, coluna in enumerate(df.columns, 1):
            aba.cell(row=1, column=col_num, value = coluna)

        #ESCREVE OS DADOS
        for row_num, row in enumerate(df.values, 2):
            for col_num, col in enumerate(row, 1):
                aba.cell(row=row_num, column=col_num, value = col)

        contador += 1

pastatrabalho.save("todas_planilhas.xlsx")



